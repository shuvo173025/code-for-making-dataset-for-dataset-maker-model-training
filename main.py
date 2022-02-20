import openpyxl

moderate = []
local = []
s = 1

wb = openpyxl.load_workbook("dataset_making.xlsx")

ws = wb['dataset_making']

colomn = ws["A"]

for i in range(2,len(colomn)+1):
    C = 'A' + str(i)
    D = 'B' + str(i)
    moderate.append(ws[C].value)
    local.append(ws[D].value)

temp = []
temp2 = []
temp3 = []

for i,x in enumerate(local):
    if '/' in x:
        temp.append(x)
        temp2.append(i)
        temp3.append(moderate[i])

for e,y in enumerate(temp):
    listB = y.split("/")
    ws[('B' + str(temp2[e] + 2))].value = listB[0]
    for j in range(1,len(listB)):
        colomn = ws["A"]
        E = 'A' + str(len(colomn)+1)
        F = 'B' + str(len(colomn)+1)
        ws[E].value = temp3[e]
        ws[F].value = listB[j]
    #wb.save("dataset_making.xlsx")